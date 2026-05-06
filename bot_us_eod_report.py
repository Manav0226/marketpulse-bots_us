from __future__ import annotations

import csv
import datetime as dt
import json
import logging
from collections import defaultdict
from pathlib import Path
from zoneinfo import ZoneInfo

from core.config_loader import US_RESEARCH_TG_CHAT, US_RESEARCH_TG_TOKEN
from core.us_market_scheduler import is_us_trading_day
from marketpulse_runtime import resolve_log_dir, resolve_report_dir, resolve_state_dir


ET = ZoneInfo("America/New_York")
UTC = dt.timezone.utc
STATE_DIR = resolve_state_dir()
LOG_DIR = resolve_log_dir()
REPORT_DIR = resolve_report_dir()
WORKBOOK_PATH = REPORT_DIR / "MarketPulse_TradeLog.xlsx"
STATE_PATH = STATE_DIR / "us_eod_report_state.json"
REPORT_STATUS_PATH = STATE_DIR / "us_report_status.json"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("us_eod_report")


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _parse_iso(value: str | None) -> dt.datetime | None:
    if not value:
        return None
    try:
        parsed = dt.datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return parsed.astimezone(UTC)
    except Exception:
        return None


def should_run_eod_report(now: dt.datetime, last_run: str | None) -> bool:
    current_utc = now.astimezone(UTC)
    if not is_us_trading_day(current_utc):
        return False
    current_et = current_utc.astimezone(ET)
    cutoff = current_et.replace(hour=16, minute=10, second=0, microsecond=0)
    if current_et < cutoff:
        return False
    previous = _parse_iso(last_run)
    if previous is None:
        return True
    return previous.astimezone(ET).date() != current_et.date()


def _load_report_state() -> dict:
    return _read_json(STATE_PATH) or {"last_run": None}


def _save_report_state(state: dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, indent=2), encoding="utf-8")


def _save_report_status(status: dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_STATUS_PATH.write_text(json.dumps(status, indent=2), encoding="utf-8")


def _load_trade_records() -> list[dict]:
    records: list[dict] = []
    for path in sorted(LOG_DIR.glob("uscrp4_trades_*.csv")):
        try:
            with path.open("r", newline="", encoding="utf-8") as handle:
                for row in csv.DictReader(handle):
                    timestamp = _parse_iso(row.get("time"))
                    if timestamp is None:
                        continue
                    pnl = float(row.get("pnl", 0) or 0)
                    qty_raw = row.get("qty", "")
                    qty = int(float(qty_raw)) if str(qty_raw).strip() else 0
                    records.append(
                        {
                            "timestamp": timestamp,
                            "date": timestamp.astimezone(ET).date().isoformat(),
                            "time": timestamp.astimezone(ET).strftime("%H:%M:%S"),
                            "market": row.get("market", ""),
                            "symbol": row.get("symbol", ""),
                            "action": row.get("action", ""),
                            "qty": qty,
                            "price": float(row.get("price", 0) or 0),
                            "sl": float(row.get("sl", 0) or 0) if str(row.get("sl", "")).strip() else None,
                            "target": float(row.get("target", 0) or 0) if str(row.get("target", "")).strip() else None,
                            "score": row.get("score", ""),
                            "reasons": row.get("reasons", ""),
                            "order_id": row.get("order_id", ""),
                            "status": row.get("status", ""),
                            "pnl": pnl,
                        }
                    )
        except Exception as exc:
            log.warning("Skipping trade log %s: %s", path.name, exc)
    return sorted(records, key=lambda item: item["timestamp"])


def _build_daily_summary(records: list[dict]) -> list[dict]:
    summary: dict[tuple[str, str], dict] = {}
    for row in records:
        key = (row["date"], row["market"])
        current = summary.setdefault(
            key,
            {"date": row["date"], "market": row["market"], "entries": 0, "exits": 0, "closed_pnl": 0.0},
        )
        if row["action"] in {"BUY", "SELL"}:
            current["entries"] += 1
        if row["action"] == "EXIT" or row["status"] == "CLOSED":
            current["exits"] += 1
            current["closed_pnl"] = round(current["closed_pnl"] + float(row["pnl"]), 2)
    return [summary[key] for key in sorted(summary.keys())]


def _write_sheet_headers(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)


def build_workbook() -> Path:
    from openpyxl import Workbook, load_workbook

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    records = _load_trade_records()
    bot_state = _read_json(STATE_DIR / "bot_state.json").get("bots", {}).get("us_v4", {})
    father = _read_json(STATE_DIR / "father_opinion.json")
    weekly = _read_json(STATE_DIR / "us_weekly_brief.json")
    supervision = _read_json(STATE_DIR / "us_supervision.json")

    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
    else:
        wb = Workbook()

    for sheet_name in ["US All Trades", "US Daily Summary", "US Open Snapshot", "US Control Tower"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb["Sheet"].max_row == 1 and wb["Sheet"].cell(1, 1).value is None:
        wb.remove(wb["Sheet"])

    trades_ws = wb.create_sheet("US All Trades", 0)
    _write_sheet_headers(
        trades_ws,
        ["Date", "Time ET", "Market", "Symbol", "Action", "Qty", "Price", "SL", "Target", "Score", "Reasons", "Order ID", "Status", "PnL"],
    )
    for row_idx, row in enumerate(records, start=2):
        trades_ws.cell(row_idx, 1, row["date"])
        trades_ws.cell(row_idx, 2, row["time"])
        trades_ws.cell(row_idx, 3, row["market"])
        trades_ws.cell(row_idx, 4, row["symbol"])
        trades_ws.cell(row_idx, 5, row["action"])
        trades_ws.cell(row_idx, 6, row["qty"])
        trades_ws.cell(row_idx, 7, row["price"])
        trades_ws.cell(row_idx, 8, row["sl"])
        trades_ws.cell(row_idx, 9, row["target"])
        trades_ws.cell(row_idx, 10, row["score"])
        trades_ws.cell(row_idx, 11, row["reasons"])
        trades_ws.cell(row_idx, 12, row["order_id"])
        trades_ws.cell(row_idx, 13, row["status"])
        trades_ws.cell(row_idx, 14, row["pnl"])

    summary_ws = wb.create_sheet("US Daily Summary")
    _write_sheet_headers(summary_ws, ["Date", "Market", "Entries", "Exits", "Closed PnL"])
    for row_idx, row in enumerate(_build_daily_summary(records), start=2):
        summary_ws.cell(row_idx, 1, row["date"])
        summary_ws.cell(row_idx, 2, row["market"])
        summary_ws.cell(row_idx, 3, row["entries"])
        summary_ws.cell(row_idx, 4, row["exits"])
        summary_ws.cell(row_idx, 5, row["closed_pnl"])

    snapshot_ws = wb.create_sheet("US Open Snapshot")
    _write_sheet_headers(snapshot_ws, ["Type", "Symbol/Market", "Venue", "Mode", "Confidence", "Risk Budget", "Opened At", "Status"])
    row_idx = 2
    for symbol, pos in sorted((bot_state.get("positions", {}) or {}).items()):
        snapshot_ws.cell(row_idx, 1, "position")
        snapshot_ws.cell(row_idx, 2, symbol)
        snapshot_ws.cell(row_idx, 3, pos.get("venue", "us_equity"))
        snapshot_ws.cell(row_idx, 4, pos.get("strategy_mode", pos.get("signal", "")))
        snapshot_ws.cell(row_idx, 5, pos.get("confidence", ""))
        snapshot_ws.cell(row_idx, 6, pos.get("risk_budget", ""))
        snapshot_ws.cell(row_idx, 7, pos.get("opened_at", ""))
        snapshot_ws.cell(row_idx, 8, "OPEN")
        row_idx += 1
    for symbol, bet in sorted((bot_state.get("bets", {}) or {}).items()):
        snapshot_ws.cell(row_idx, 1, "bet")
        snapshot_ws.cell(row_idx, 2, symbol)
        snapshot_ws.cell(row_idx, 3, bet.get("venue", "polymarket"))
        snapshot_ws.cell(row_idx, 4, bet.get("strategy_mode", "event"))
        snapshot_ws.cell(row_idx, 5, bet.get("confidence", ""))
        snapshot_ws.cell(row_idx, 6, bet.get("risk_budget", ""))
        snapshot_ws.cell(row_idx, 7, bet.get("opened_at", ""))
        snapshot_ws.cell(row_idx, 8, "OPEN")
        row_idx += 1

    control_ws = wb.create_sheet("US Control Tower")
    control_rows = [
        ("generated_at", dt.datetime.now(UTC).isoformat()),
        ("safe_mode", json.dumps(bot_state.get("safe_mode", {}), default=str)),
        ("health", json.dumps(bot_state.get("health", {}), default=str)),
        ("performance", json.dumps(bot_state.get("performance", {}), default=str)),
        ("promotion_status", json.dumps(bot_state.get("promotion_status", {}), default=str)),
        ("father_us_mode", father.get("us", {}).get("mode", "")),
        ("market_regime", father.get("market_regime", "")),
        ("weekly_candidates", ", ".join(item.get("symbol", "") for item in (weekly.get("weekly_candidates", []) or [])[:8])),
        ("blocked_symbols", ", ".join(supervision.get("blocked_symbols", []) or [])),
        ("event_risk_symbols", ", ".join(supervision.get("event_risk_symbols", []) or [])),
        ("source_warnings", ", ".join(supervision.get("source_warnings", []) or [])),
    ]
    _write_sheet_headers(control_ws, ["Field", "Value"])
    for row_idx, (field, value) in enumerate(control_rows, start=2):
        control_ws.cell(row_idx, 1, field)
        control_ws.cell(row_idx, 2, value)

    wb.save(WORKBOOK_PATH)
    wb.close()
    return WORKBOOK_PATH


def run_if_due(now: dt.datetime | None = None) -> bool:
    current = now or dt.datetime.now(UTC)
    state = _load_report_state()
    if not should_run_eod_report(current, state.get("last_run")):
        return False
    path = build_workbook()
    state["last_run"] = current.isoformat()
    _save_report_state(state)
    try:
        from notifier import Notifier

        notifier = Notifier(US_RESEARCH_TG_TOKEN, US_RESEARCH_TG_CHAT)
        notifier.alert(
            f"US EOD workbook refreshed\nDate ET: {current.astimezone(ET).date().isoformat()}\nPath: {path}"
        )
        notifier.send_telegram_document(
            path,
            caption=f"US EOD workbook {current.astimezone(ET).date().isoformat()}",
            silent=True,
        )
    except Exception:
        pass
    _save_report_status(
        {
            "generated_at": current.isoformat(),
            "workbook_path": str(path),
            "date_et": current.astimezone(ET).date().isoformat(),
            "state_path": str(STATE_PATH),
            "sent_to_telegram": bool(US_RESEARCH_TG_TOKEN and US_RESEARCH_TG_CHAT),
        }
    )
    return True


def main() -> None:
    ran = run_if_due()
    if ran:
        log.info("US EOD workbook updated: %s", WORKBOOK_PATH)
    else:
        log.info("US EOD workbook not due")


if __name__ == "__main__":
    main()
